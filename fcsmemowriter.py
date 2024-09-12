#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import openai
import boto3
from openpyxl import load_workbook
from PIL import Image
import io
import docx
import os
import time
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)

# Set up configuration
OPEN_AI_API_KEY = os.getenv("OPEN_AI_API_KEY")
ASSISTANT_ID = os.getenv("OPENAI_ASSISTANT_ID")
AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_KEY")
AWS_REGION = os.getenv("AWS_REGION")
S3_BUCKET_NAME = os.getenv("S3_BUCKET_NAME")

# Initialize OpenAI client
client = openai.Client(api_key=OPEN_AI_API_KEY)

# Initialize AWS S3 client
s3 = boto3.client('s3', 
                  aws_access_key_id=AWS_ACCESS_KEY, 
                  aws_secret_access_key=AWS_SECRET_KEY,
                  region_name=AWS_REGION)

def process_excel(file, sheet_name=None):
    if file is None:
        logging.warning("No Excel file provided")
        return None
    try:
        logging.info(f"Processing Excel file: {file.name}")
        workbook = load_workbook(file)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        logging.info("Excel file processed successfully")
        return data
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        st.error(f"Error processing Excel file: {str(e)}")
        return None

def process_supplemental(file):
    if file is None:
        return ""
    try:
        if file.type == "application/pdf":
            # You might want to implement PDF processing here if needed
            return "PDF content placeholder"
        elif file.type in ["image/jpeg", "image/jpg"]:
            image = Image.open(file)
            # Placeholder for image processing - you might want to implement OCR here
            return "Image content placeholder"
        else:
            return "Unsupported file type"
    except Exception as e:
        logging.error(f"Error processing supplemental file: {str(e)}")
        st.error(f"Error processing supplemental file: {str(e)}")
        return None

# ... (keep other functions like generate_memo, create_word_document, save_to_s3 as they were) ...

def update_stage():
    if 'term_sheet' in st.session_state and 'pricing_table' in st.session_state:
        if st.session_state.term_sheet is not None and st.session_state.pricing_table is not None:
            st.session_state.stage = 'ready'
        else:
            st.session_state.stage = 'upload'

def reset_state():
    st.session_state.stage = 'initial'
    if 'term_sheet' in st.session_state:
        del st.session_state.term_sheet
    if 'pricing_table' in st.session_state:
        del st.session_state.pricing_table
    if 'supplemental' in st.session_state:
        del st.session_state.supplemental

def main():
    st.title("AI Memo Writer")
    
    # Initialize session state
    if 'stage' not in st.session_state:
        st.session_state.stage = 'initial'
    
    if st.session_state.stage == 'initial':
        if st.button("Commence"):
            st.session_state.stage = 'upload'
    
    if st.session_state.stage in ['upload', 'ready']:
        st.file_uploader("Upload Term Sheet (XLSX)", type="xlsx", key="term_sheet", on_change=update_stage)
        st.file_uploader("Upload Pricing Table (XLSX)", type="xlsx", key="pricing_table", on_change=update_stage)
        st.file_uploader("Upload Supplemental Document (Optional)", type=["pdf", "jpg", "jpeg"], key="supplemental")
        
        if st.session_state.stage == 'ready':
            if st.button("Generate Memo"):
                st.session_state.stage = 'generate'
    
    if st.session_state.stage == 'generate':
        with st.spinner("Processing files and generating memo..."):
            if 'term_sheet' not in st.session_state or st.session_state.term_sheet is None:
                st.error("Term sheet is missing. Please upload it and try again.")
                st.session_state.stage = 'upload'
                return

            if 'pricing_table' not in st.session_state or st.session_state.pricing_table is None:
                st.error("Pricing table is missing. Please upload it and try again.")
                st.session_state.stage = 'upload'
                return

            term_sheet_data = process_excel(st.session_state.term_sheet)
            pricing_data = process_excel(st.session_state.pricing_table)
            supplemental_text = process_supplemental(st.session_state.supplemental if 'supplemental' in st.session_state else None)
            
            if term_sheet_data and pricing_data:
                # Convert Excel data to string format for memo generation
                term_sheet_text = "\n".join([", ".join(map(str, row)) for row in term_sheet_data])
                pricing_text = "\n".join([", ".join(map(str, row)) for row in pricing_data])
                
                memo_text = generate_memo(term_sheet_text, pricing_text, supplemental_text)
                
                if memo_text:
                    doc_bytes = create_word_document(memo_text)
                    
                    if doc_bytes:
                        s3_link = save_to_s3(doc_bytes)
                        
                        if s3_link:
                            st.success("Memo generated successfully!")
                            st.markdown(f"[Download Memo]({s3_link})")
                        else:
                            st.error("Failed to save memo to S3.")
                    else:
                        st.error("Failed to create Word document.")
                else:
                    st.error("Failed to generate memo.")
            else:
                st.error("Failed to process input files.")
        
        if st.button("Start Over"):
            reset_state()

if __name__ == "__main__":
    main()

